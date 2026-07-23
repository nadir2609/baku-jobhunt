"""Build baku_it_companies.xlsx from companies.json + send_log.json.

Safe to re-run at any time; it never touches the network.
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
COMPANIES = ROOT / "companies_final.json"
SENT_YESTERDAY = ROOT / "sent_yesterday.json"
LOG = ROOT / "send_log.json"
OUT = ROOT / "baku_it_companies.xlsx"

HEADERS = [
    "Name",
    "Main Email",
    "HR Email",
    "Emailed Today (yes/no) and to which(or both)",
    "Already Emailed Yesterday (yes/no)",
]


def load(path, default):
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return default


def main():
    companies = load(COMPANIES, [])
    log = load(LOG, [])
    yesterday = load(SENT_YESTERDAY, {})

    skip_addrs = {a.lower() for a in yesterday.get("addresses", [])}
    skip_domains = {d.lower() for d in yesterday.get("domains", [])}

    # address -> status, from the send log
    sent_ok = {e["to"].lower() for e in log if e.get("status") == "sent"}
    failures = {
        e["to"].lower(): e.get("error", "error")
        for e in log
        if e.get("status") == "failed"
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Baku IT Companies"
    ws.append(HEADERS)

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2F5597")
    for cell in ws[1]:
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for c in companies:
        name = c.get("name", "?")
        main = c.get("main_email") or "0"
        hr = c.get("hr_email") or "0"

        addrs = [a for a in (main, hr) if a != "0"]
        doms = {a.split("@")[-1].lower() for a in addrs}
        was_yesterday = bool(
            doms & skip_domains or {a.lower() for a in addrs} & skip_addrs
        )

        hit = []
        if main != "0" and main.lower() in sent_ok:
            hit.append("main")
        if hr != "0" and hr.lower() in sent_ok:
            hit.append("HR")

        if len(hit) == 2:
            today = "yes - both"
        elif hit:
            today = f"yes - {hit[0]}"
        else:
            errs = [failures[a.lower()] for a in addrs if a.lower() in failures]
            if errs:
                today = f"no - send failed: {errs[0][:120]}"
            elif was_yesterday:
                today = "no - emailed yesterday"
            elif not addrs:
                today = "no - no email found"
            else:
                today = "no"

        ws.append([name, main, hr, today, "yes" if was_yesterday else "no"])

    widths = [38, 34, 34, 46, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(OUT)

    total = len(companies)
    with_any = sum(
        1
        for c in companies
        if (c.get("main_email") or "0") != "0" or (c.get("hr_email") or "0") != "0"
    )
    print(f"Wrote {OUT}")
    print(f"  companies:            {total}")
    print(f"  with >=1 email:       {with_any}")
    print(f"  addresses emailed:    {len(sent_ok)}")
    print(f"  send failures:        {len(failures)}")


if __name__ == "__main__":
    main()
